import requests # for making standard html requests
from bs4 import BeautifulSoup # magical tool for parsing html data
import time
import pandas as pd
from openpyxl import load_workbook
from datetime import date
from openpyxl import Workbook
import os
import smtplib
from email.mime.text import MIMEText
import json # for parsing data
from pandas import DataFrame as df # premier library for data organization

#my custom_made modules
from Utils import floatRepr, formatTitle, getCredentials, getMinRowValue
from product_class import product_class

def readProductLists(fileList):
    productList = []
    try:
        with open(fileList, 'r') as f:
            for line in f.readlines():
                line = line.strip()
                if (line == ""):
                    continue
                URL, perc, email_recips = line.split(',')
                recips = email_recips.split()
                tempProduct = product_class(URL, perc, recips)

                tempProduct.print_attrs()
                productList.append(tempProduct)
    except Exception as e:
        print(e)
    return productList

def request2BfSoupObj(root_url, url_path):
    page = requests.get("{}{}".format(root_url,url_path))
    print("\nRequesting Page URL: {}{}\n".format(root_url,url_path))
    if page.status_code == 200: print("\nRequest OK: Status code {}\n".format(page.status_code))
    else:
        print("\nError with the request:response: {}\n".format(page.status_code))
        raise ConnectionError("\nError with the request:response: {}\n".format(page.status_code))
        return 0
    page.encoding = 'ISO-885901'
    soup = BeautifulSoup(page.text, 'html.parser')      #using the html parser, easier to search in browser
    return soup

def productPrice(soup, product_title):
    #get main-container-inner
    productPriceTag = soup.body.find("div", class_="main-container-inner")
    #go to product-page-pricing
    productHighlight = productPriceTag.find("div", class_="product-page-pricing")
    # extract 'new-price' -> so-called reduced one in some cases
    productNewPrice = productHighlight.find("p", class_="product-new-price")

    # extract main price, it is the pos 0 in the list of contents
    mainNewPrice = productNewPrice.contents[0].strip()

    # extract the secondary price, it is under an <sup> tag
    secNewPrice = productNewPrice.sup.string

    # extract 'old-price' -> so-called unreduced one
    # careful how we treat it, it may exist => extract, it may not exist => None => ==new-price
    if (productHighlight.find("p", class_="product-old-price") is None):
        mainOldPrice = mainNewPrice
        secOldPrice  = secNewPrice
    else:
        productOldPrice = productHighlight.find("p", class_="product-old-price")
        mainOldPrice = productOldPrice.s.contents[0]
        secOldPrice  = productOldPrice.s.sup.string


    if (mainNewPrice is not None and secNewPrice is not None) and (mainOldPrice is not None and secOldPrice is not None):
        formatedProudctBasePrice    = "{},{}".format(mainOldPrice, secOldPrice)
        formatedProudctReducedPrice = "{},{}".format(mainNewPrice, secNewPrice)
        return formatedProudctBasePrice, formatedProudctReducedPrice
    else:
        raise ValueError("\nPrice could not be extracted for product {}\n".format(product_title))
        print(productHighlight)
        print(mainNewPrice, secNewPrice)
        print(mainOldPrice, secOldPrice)
        return 0

def email_nofifier(usr, pswd, recip, body, subject):
    #first customize the email
    user = usr
    password = pswd
    #sent_to is a list, we need it in a string format of recip1, recip2, etc...
    #subject = "Test email"
    email = MIMEText("{}".format(body))
    sender = usr
    recipients = recip
    email['Subject'] = subject
    email['From'] = sender
    email['To'] = ", ".join(recipients)


    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
        server.ehlo()
        server.login(user, password)
        server.sendmail(sender, recipients, email.as_string())
        server.close()

        print('Email sent!')
    except Exception as e:
        print("Something went wrong!")
        print(e)

def main():
    #load the list of products
    #opening spreadsheet
    excelName = "Prices.xlsx"
    today = date.today()
    today = today.strftime("%B %d, %Y")
    bot_user, bot_pswd = getCredentials()
    #email_nofifier(bot_user, bot_pswd, ["florin.firanescu@gmail.com"])
    #print(calculate_percDecrease(0, 89))
    #exit()
    if not (os.path.exists(excelName)):
        print("Creating a new excel file named : {}".format(excelName))
        workbook = Workbook()
        sheet = workbook.active

        workbook.save(filename=excelName)

    excel_file = load_workbook(filename=excelName)

    productList = readProductLists("Products.txt")

    if len(productList) == 0:
        exit("Product list is empty. Exiting...")
    for pageProduct in productList:
        soup = BeautifulSoup()
        pageRoot    = "https://www.emag.ro/"
        URL = pageRoot + pageProduct.URL


        try:
            soup = request2BfSoupObj(pageRoot, pageProduct.URL)
        except Exception as e:
            print(e)
            print("Bad request response. Next product!")
            continue
        try:
            title = soup.title.string
            basePrice, reducedPrice = productPrice(soup, title)
            pageProduct.actual_basePrice    = basePrice
            pageProduct.actual_reducedPrice = reducedPrice
            print("URL                  : {}".format(URL))
            print("Product name         : {}\nProduct base price   : {}\nProduct reduced price: {}".format(title, pageProduct.actual_basePrice, pageProduct.actual_reducedPrice))
        except Exception as e:
            print(e)
            print("Continuing to next product")
            continue

        true_title = title
        title = formatTitle(title)
        recips = ', '.join(pageProduct.email_recips)
        if title not in excel_file.sheetnames:
            excel_file.create_sheet(title)
            excel_file[title].append(['Date', 'Link', 'BasePrice', 'ReducedPrice', 'Email_recip', 'Email_notification'])
            # first attempt
            excel_file[title].append([today, URL, pageProduct.actual_basePrice, pageProduct.actual_reducedPrice, recips,
                                 pageProduct.email_triggered])
            excel_file.save(excelName)
            continue
        productSheet = excel_file[title]
        pageProduct.prev_price = productSheet['D2'].value

        if productSheet['A{}'.format(productSheet.max_row)].value == today:
            print("Same day")
        else:
            if (float(pageProduct.calculatePercentage()) > float(pageProduct.percentage)
                    and getMinRowValue(productSheet, 'D') > floatRepr(pageProduct.actual_reducedPrice)):
                email_nofifier(bot_user, bot_pswd, pageProduct.email_recips, pageProduct.getBody(true_title, URL), pageProduct.getSubject())
                pageProduct.email_triggered = 1
            productSheet.append([today, URL, pageProduct.actual_basePrice, pageProduct.actual_reducedPrice, recips, pageProduct.email_triggered ])
        time.sleep(2)
    excel_file.save(excelName)

if __name__ == "__main__":
    main()